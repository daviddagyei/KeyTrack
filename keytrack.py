import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

workbook = openpyxl.load_workbook('key_distribution.xlsx')
sheet = workbook.active

def get_current_data():
    keys_distribution = {}
    for row in range(2, sheet.max_row + 1):
        room = sheet.cell(row=row, column=1).value
        available_keys = sheet.cell(row=row, column=2).value

        def parse_list(cell_value):
            return cell_value.split(', ') if cell_value else []

        collected_by = parse_list(sheet.cell(row=row, column=3).value)
        lost_keys = parse_list(sheet.cell(row=row, column=4).value)
        borrowed_spare_keys = parse_list(sheet.cell(row=row, column=5).value)
        returned_keys = parse_list(sheet.cell(row=row, column=6).value)

        keys_distribution[room] = {
            "available_keys": available_keys,
            "collected_by": collected_by,
            "lost_keys": lost_keys,
            "borrowed_spare_keys": borrowed_spare_keys,
            "returned_keys": returned_keys
        }
    return keys_distribution

def update_excel(data):
    for row in range(2, sheet.max_row + 1):
        room = sheet.cell(row=row, column=1).value
        if room in data:
            sheet.cell(row=row, column=2, value=data[room]["available_keys"])
            sheet.cell(row=row, column=3, value=", ".join(data[room]["collected_by"]))
            sheet.cell(row=row, column=4, value=", ".join(data[room]["lost_keys"]))
            sheet.cell(row=row, column=5, value=", ".join(data[room]["borrowed_spare_keys"]))
            sheet.cell(row=row, column=6, value=", ".join(data[room]["returned_keys"]))
    workbook.save('key_distribution.xlsx')

def get_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def handle_action(task, room, student):
    data = get_current_data()
    room = room.strip()
    
    if room not in data:
        messagebox.showerror("Error", "Invalid room number.")
        return

    student = student.strip()
    timestamped_name = f"{student} ({get_timestamp()})"

    if task == "collect":
        if data[room]["available_keys"] > 0:
            data[room]["available_keys"] -= 1
            data[room]["collected_by"].append(timestamped_name)
            messagebox.showinfo("Success", f"{student} collected a key for {room}.")
        else:
            messagebox.showerror("Error", "No more keys available for this room.")
    elif task == "return":
        match = next((name for name in data[room]["collected_by"] if student in name), None)
        if match:
            data[room]["collected_by"].remove(match)
            data[room]["available_keys"] += 1
            return_timestamped_name = f"{student} ({get_timestamp()})"
            data[room]["returned_keys"].append(return_timestamped_name)
            messagebox.showinfo("Success", f"{student} returned a key for {room}.")
        else:
            messagebox.showerror("Error", "No record of key collection for this student.")
    elif task == "report_lost":
        match = next((name for name in data[room]["collected_by"] if student in name), None)
        if match:
            data[room]["collected_by"].remove(match)
            data[room]["lost_keys"].append(timestamped_name)
            messagebox.showinfo("Success", f"{student} reported a lost key for {room}.")
        else:
            messagebox.showerror("Error", "You have not collected a key for this room.")
    elif task == "borrow_spare":
        if data[room]["available_keys"] > 0:
            data[room]["available_keys"] -= 1
            data[room]["borrowed_spare_keys"].append(timestamped_name)
            messagebox.showinfo("Success", f"{student} borrowed a spare key for {room}.")
        else:
            messagebox.showerror("Error", "No spare keys available for this room.")

    update_excel(data)

def on_action_select(task):
    room = room_entry.get()
    student = student_entry.get()
    handle_action(task, room, student)


root = tk.Tk()
root.title("Key Management System")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(padx=10, pady=10)

tk.Label(frame, text="Room Number:").grid(row=0, column=0, sticky="e")
room_entry = tk.Entry(frame)
room_entry.grid(row=0, column=1)

tk.Label(frame, text="Student Name:").grid(row=1, column=0, sticky="e")
student_entry = tk.Entry(frame)
student_entry.grid(row=1, column=1)

actions = ["collect", "return", "report_lost", "borrow_spare"]
for i, action in enumerate(actions, start=2):
    btn = tk.Button(frame, text=action.capitalize(), command=lambda a=action: on_action_select(a))
    btn.grid(row=i, column=0, columnspan=2, sticky="ew")

root.mainloop()
